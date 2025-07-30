from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from pathlib import Path
import sys


def create_main_sheet(ws, items):
    """Create the main sheet with items and their details."""
    ws.append(
        [
            "id",
            "item",
            "quantity",
            "prix /u",
            "Prix total",
            "Stock",
            "Qty restante",
            "Cout restant",
        ]
    )

    for i, item in enumerate(items, start=1):
        row_idx = i + 1
        ws.append(
            [
                i,
                item[0],
                item[1],
                1,
                f"=C{row_idx}*D{row_idx}",
                0,
                f"=C{row_idx}-F{row_idx}",
                f"=G{row_idx}*D{row_idx}",
            ]
        )

    # Ajout des totaux
    summary_col = ws.max_column + 2
    summary_col_letter = get_column_letter(summary_col)
    summary_col2_letter = get_column_letter(summary_col + 1)
    ws[f"{summary_col_letter}1"] = "Cout Total"
    ws[f"{summary_col2_letter}1"] = "=SUBTOTAL(109,TableauItems[Prix total])"
    ws[f"{summary_col_letter}2"] = "Cout Restant"
    ws[f"{summary_col2_letter}2"] = "=SUBTOTAL(109,TableauItems[Cout restant])"
    table_ref = f"A1:H{ws.max_row}"
    tab = Table(displayName="TableauItems", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)


def create_equipement_sheet(wb, data_equipements):
    """Create sheets for each equipment with its components."""
    for nom_equipement, compos in data_equipements.items():
        ws_equip = wb.create_sheet(title=nom_equipement)
        ws_equip.append(["Composant", "Quantité", "Prix /u", "Cout Total"])
        for idx, compo in enumerate(compos, start=2):
            # Extraction du nom et de la quantité du composant
            # Ajout des formules pour le prix unitaire et le coût total
            prix_u_formula = f'=IFERROR(VLOOKUP(A{idx},Items!B:D,3,FALSE),"")'
            cout_total_formula = f"=B{idx}*C{idx}"
            ws_equip.append(
                [compo[0], int(compo[1]), prix_u_formula, cout_total_formula]
            )
        table_ref = f"A1:D{len(compos)+1}"
        tab = Table(
            displayName=f"Tableau_{nom_equipement.replace(' ', '_').replace('\\', '_').replace('/', '_')}",
            ref=table_ref,
        )
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        ws_equip.add_table(tab)
        ws_equip["F1"] = "Cout Total"
        ws_equip["G1"] = (
            f'=SUBTOTAL(109,Tableau_{nom_equipement.replace(" ", "_").replace("\\", "_").replace("/", "_")}[Cout Total])'
        )


def init_excel():
    filename= ''
    while not filename:
        filename = input("Entrez le nom du fichier excel de destination : ").strip()
        if not filename:
            print("Le nom du fichier ne peut pas être vide. Veuillez réessayer.")

    base_path = (
        Path(sys.executable).parent
        if getattr(sys, "frozen", False)
        else Path(__file__).parent
    )
    fileresult = base_path / "result" / f"{filename}.xlsx"
    if not fileresult.parent.exists():
        fileresult.parent.mkdir(parents=True, exist_ok=True)

    """Initialize the Excel workbook and create the main sheet."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Items"

    return wb, ws, base_path, fileresult


def save_excel(wb, fileresult):
    """Save the Excel workbook to the specified file."""
    wb.save(fileresult)
    infos = [
        f"Fichier Excel créé avec succès : {fileresult}",
        # f"Nombre d'items traités : {len(items)}",
        # f"Nombre d'équipements traités : {len(equipements)}",
        f"Nombre de feuilles créées : {len(wb.sheetnames)}",
        f"Feuilles créées : {', '.join(wb.sheetnames)}",
        "Fin du traitement.",
        "Le fichier Excel est prêt à être utilisé.",
        "Vous pouvez l'ouvrir avec Excel ou un autre tableur compatible.",
        "Merci d'avoir utilisé ce script !",
    ]
    print("\n".join(infos))
    input("\nAppuyez sur Entrée pour quitter...")
